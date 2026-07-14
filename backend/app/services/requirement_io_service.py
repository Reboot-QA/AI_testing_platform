import json
import re
import uuid
import zipfile
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy.orm import Session, joinedload

from app.models.project import Project
from app.models.requirement import Requirement
from app.models.user import User

TYPE_LABELS = {
    "functional": "功能",
    "api": "接口",
    "performance": "性能",
    "security": "安全",
}

TYPE_ALIASES = {
    "功能": "functional",
    "功能测试": "functional",
    "functional": "functional",
    "接口": "api",
    "接口测试": "api",
    "api": "api",
    "性能": "performance",
    "性能测试": "performance",
    "performance": "performance",
    "安全": "security",
    "安全测试": "security",
    "security": "security",
}

STATUS_LABELS = {
    "draft": "草稿",
    "approved": "已评审",
    "closed": "已关闭",
}

STATUS_ALIASES = {
    "草稿": "draft",
    "draft": "draft",
    "已评审": "approved",
    "approved": "approved",
    "已关闭": "closed",
    "closed": "closed",
}

SOURCE_LABELS = {
    "manual": "手动",
    "ai_document": "文档解析",
}

SOURCE_ALIASES = {
    "手动": "manual",
    "manual": "manual",
    "文档解析": "ai_document",
    "ai_document": "ai_document",
}

ALLOWED_TYPES = set(TYPE_LABELS)
ALLOWED_PRIORITIES = {"P0", "P1", "P2", "P3"}
ALLOWED_STATUSES = set(STATUS_LABELS)
ALLOWED_SOURCES = set(SOURCE_LABELS)

EXCEL_HEADERS = ["ID", "标题", "类型", "优先级", "状态", "来源", "描述"]
HEADER_ALIASES = {
    "id": "ID",
    "标题": "标题",
    "title": "标题",
    "类型": "类型",
    "type": "类型",
    "req_type": "类型",
    "优先级": "优先级",
    "priority": "优先级",
    "状态": "状态",
    "status": "状态",
    "来源": "来源",
    "source": "来源",
    "描述": "描述",
    "description": "描述",
}

GROUP_SUFFIX = re.compile(r"（\d+）$")


def build_content_disposition(filename: str, fallback: str) -> str:
    encoded = quote(filename, safe="")
    return f'attachment; filename="{fallback}"; filename*=UTF-8\'\'{encoded}'


def normalize_type(value: Optional[str]) -> str:
    text = (value or "").strip()
    if not text:
        return "functional"
    mapped = TYPE_ALIASES.get(text) or TYPE_ALIASES.get(text.lower())
    if mapped in ALLOWED_TYPES:
        return mapped
    raise ValueError(f"无效类型: {value}")


def normalize_priority(value: Optional[str]) -> str:
    text = (value or "P1").strip().upper()
    if text in ALLOWED_PRIORITIES:
        return text
    raise ValueError(f"无效优先级: {value}")


def normalize_status(value: Optional[str]) -> str:
    text = (value or "draft").strip()
    mapped = STATUS_ALIASES.get(text) or STATUS_ALIASES.get(text.lower())
    if mapped in ALLOWED_STATUSES:
        return mapped
    raise ValueError(f"无效状态: {value}")


def normalize_source(value: Optional[str]) -> str:
    text = (value or "manual").strip()
    mapped = SOURCE_ALIASES.get(text) or SOURCE_ALIASES.get(text.lower())
    if mapped in ALLOWED_SOURCES:
        return mapped
    raise ValueError(f"无效来源: {value}")


def _strip_group_label(label: str) -> str:
    return GROUP_SUFFIX.sub("", (label or "").strip()).strip()


def _infer_type_priority(ancestors: List[str]) -> Tuple[str, str]:
    req_type = "functional"
    priority = "P1"
    for label in ancestors:
        clean = _strip_group_label(label)
        if clean in TYPE_ALIASES:
            req_type = normalize_type(clean)
        elif clean.upper() in ALLOWED_PRIORITIES:
            priority = normalize_priority(clean)
        elif clean.startswith("P") and clean[:2] in ALLOWED_PRIORITIES:
            priority = normalize_priority(clean[:2])
    return req_type, priority


def _topic_id() -> str:
    return uuid.uuid4().hex


def _build_xmind_topic(
    title: str,
    children: Optional[List[Dict]] = None,
    notes: Optional[str] = None,
    *,
    structure_class: Optional[str] = None,
) -> Dict:
    topic: Dict[str, Any] = {
        "id": _topic_id(),
        "class": "topic",
        "title": title,
    }
    if structure_class:
        topic["structureClass"] = structure_class
    if notes:
        topic["notes"] = {
            "plain": {"content": notes},
            "realHTML": {"content": f"<div>{notes}</div>"},
        }
    if children:
        topic["children"] = {"attached": children}
    return topic


# 1x1 PNG，XMind 缩略图占位
_XMIND_THUMBNAIL_PNG = bytes.fromhex(
    "89504e470d0a1a0a0000000d4948445200000001000000010802000000907753de"
    "0000000c49444154789c63f80f00000101000518d4630000000049454e44ae426082"
)


def _group_requirements(requirements: List[Requirement]) -> Dict[str, Dict[str, List[Requirement]]]:
    grouped: Dict[str, Dict[str, List[Requirement]]] = {}
    for req in requirements:
        type_key = req.req_type or "functional"
        priority = req.priority or "P1"
        grouped.setdefault(type_key, {}).setdefault(priority, []).append(req)
    return grouped


def export_requirements_excel(project: Project, requirements: List[Requirement]) -> Tuple[BytesIO, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "需求点"
    ws.append(EXCEL_HEADERS)
    for col in range(1, len(EXCEL_HEADERS) + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for req in requirements:
        ws.append(
            [
                req.id,
                req.title,
                TYPE_LABELS.get(req.req_type, req.req_type),
                req.priority,
                STATUS_LABELS.get(req.status, req.status),
                SOURCE_LABELS.get(req.source, req.source),
                req.description or "",
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 48

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{project.name}_requirements.xlsx"
    return buffer, filename


def export_requirements_xmind(project: Project, requirements: List[Requirement]) -> Tuple[BytesIO, str]:
    grouped = _group_requirements(requirements)
    type_nodes: List[Dict] = []

    for type_key in ["functional", "api", "performance", "security"]:
        priority_map = grouped.get(type_key)
        if not priority_map:
            continue
        priority_nodes: List[Dict] = []
        for priority in ["P0", "P1", "P2", "P3"]:
            items = priority_map.get(priority) or []
            if not items:
                continue
            leaf_nodes = [
                _build_xmind_topic(req.title, notes=req.description or None)
                for req in sorted(items, key=lambda item: item.id, reverse=True)
            ]
            priority_nodes.append(
                _build_xmind_topic(f"{priority}（{len(items)}）", children=leaf_nodes)
            )
        type_label = TYPE_LABELS.get(type_key, type_key)
        total = sum(len(v) for v in priority_map.values())
        type_nodes.append(
            _build_xmind_topic(f"{type_label}（{total}）", children=priority_nodes)
        )

    root = _build_xmind_topic(
        f"{project.name}（{len(requirements)}）",
        children=type_nodes or None,
        structure_class="org.xmind.ui.logic.right",
    )
    sheet_id = _topic_id()
    sheet = {
        "id": sheet_id,
        "class": "sheet",
        "title": "需求点",
        "rootTopic": root,
        "theme": {
            "map": {
                "id": _topic_id(),
                "properties": {
                    "svg:fill": "#ffffff",
                    "multi-line-colors": "",
                    "color-list": "",
                },
            }
        },
    }
    metadata = {
        "dataStructureVersion": "2",
        "layoutEngineVersion": "3",
        "creator": {"name": "AI质量平台", "version": "1.0.0"},
        "activeSheetId": sheet_id,
        "modifier": "",
    }
    manifest = {
        "file-entries": {
            "content.json": {},
            "metadata.json": {},
            "Thumbnails/thumbnail.png": {},
        }
    }

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("content.json", json.dumps([sheet], ensure_ascii=False))
        archive.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False))
        archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
        archive.writestr("Thumbnails/thumbnail.png", _XMIND_THUMBNAIL_PNG)
    buffer.seek(0)
    filename = f"{project.name}_requirements.xmind"
    return buffer, filename


def _parse_excel_rows(file_bytes: bytes) -> List[Dict[str, str]]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 文件为空")

    header_row = [str(cell or "").strip() for cell in rows[0]]
    mapping: Dict[int, str] = {}
    for index, header in enumerate(header_row):
        normalized = HEADER_ALIASES.get(header) or HEADER_ALIASES.get(header.lower())
        if normalized:
            mapping[index] = normalized

    if "标题" not in mapping.values():
        raise ValueError("Excel 缺少「标题」列")

    items: List[Dict[str, str]] = []
    for row in rows[1:]:
        if not row:
            continue
        record: Dict[str, str] = {}
        for index, field in mapping.items():
            value = row[index] if index < len(row) else None
            record[field] = str(value).strip() if value is not None else ""
        title = record.get("标题", "").strip()
        if not title:
            continue
        items.append(record)
    if not items:
        raise ValueError("Excel 中没有可导入的需求行")
    return items


def _load_xmind_topics(file_bytes: bytes) -> List[Dict]:
    try:
        with zipfile.ZipFile(BytesIO(file_bytes)) as archive:
            if "content.json" in archive.namelist():
                payload = json.loads(archive.read("content.json").decode("utf-8"))
                if isinstance(payload, list):
                    topics = []
                    for sheet in payload:
                        root = sheet.get("rootTopic")
                        if root:
                            topics.append(root)
                    return topics
                if isinstance(payload, dict) and payload.get("rootTopic"):
                    return [payload["rootTopic"]]
            raise ValueError("XMind 文件缺少 content.json")
    except zipfile.BadZipFile as exc:
        raise ValueError("无效的 XMind 文件") from exc


def _collect_xmind_requirements(topic: Dict, ancestors: Optional[List[str]] = None) -> List[Dict[str, str]]:
    ancestors = list(ancestors or [])
    title = (topic.get("title") or "").strip()
    children = (topic.get("children") or {}).get("attached") or []
    notes = ((topic.get("notes") or {}).get("plain") or {}).get("content") or ""
    notes = str(notes).strip()

    items: List[Dict[str, str]] = []
    if not children:
        if not title or GROUP_SUFFIX.match(title):
            return items
        if len(ancestors) == 0:
            return items
        req_type, priority = _infer_type_priority(ancestors)
        items.append(
            {
                "标题": title,
                "描述": notes,
                "类型": TYPE_LABELS.get(req_type, req_type),
                "优先级": priority,
                "状态": "草稿",
                "来源": "手动",
            }
        )
        return items

    next_ancestors = ancestors + [title] if title else ancestors
    for child in children:
        items.extend(_collect_xmind_requirements(child, next_ancestors))
    return items


def parse_requirement_import_file(filename: str, file_bytes: bytes) -> List[Dict[str, str]]:
    lower_name = (filename or "").lower()
    if lower_name.endswith((".xlsx", ".xlsm", ".xltx")):
        return _parse_excel_rows(file_bytes)
    if lower_name.endswith(".xmind"):
        topics = _load_xmind_topics(file_bytes)
        items: List[Dict[str, str]] = []
        for topic in topics:
            items.extend(_collect_xmind_requirements(topic, []))
        if not items:
            raise ValueError("XMind 中未找到可导入的需求节点")
        return items
    raise ValueError("仅支持 .xlsx 或 .xmind 文件")


def import_requirements_from_rows(
    db: Session,
    project: Project,
    rows: List[Dict[str, str]],
    current_user: User,
) -> Tuple[int, int]:
    imported = 0
    skipped = 0
    for row in rows:
        title = row.get("标题", "").strip()
        if not title:
            skipped += 1
            continue
        try:
            req = Requirement(
                project_id=project.id,
                title=title,
                description=row.get("描述") or None,
                req_type=normalize_type(row.get("类型")),
                priority=normalize_priority(row.get("优先级")),
                status=normalize_status(row.get("状态")),
                source=normalize_source(row.get("来源")),
                created_by_id=current_user.id,
            )
        except ValueError:
            skipped += 1
            continue
        db.add(req)
        imported += 1
    if imported:
        db.commit()
    return imported, skipped


def list_project_requirements(db: Session, project_id: int) -> List[Requirement]:
    return (
        db.query(Requirement)
        .options(joinedload(Requirement.creator), joinedload(Requirement.project))
        .filter(Requirement.project_id == project_id)
        .order_by(Requirement.id.desc())
        .all()
    )
