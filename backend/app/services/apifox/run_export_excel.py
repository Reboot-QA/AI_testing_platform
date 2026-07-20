"""Apifox 运行报告导出 · Excel。"""

from io import BytesIO
from typing import Any, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.routers.apifox.run_schemas import RunOut
from app.services.apifox.run_export_common import (
    STATUS_LABELS,
    _format_assertions,
    _format_dt,
    _format_duration,
    _format_report_time,
    _pretty_json_text,
)


def _autosize_row(ws, row_num: int, min_height: int = 18, max_height: int = 120) -> None:
    max_lines = 1
    for cell in ws[row_num]:
        if cell.value is None:
            continue
        text = str(cell.value)
        col_width = ws.column_dimensions[get_column_letter(cell.column)].width or 10
        chars_per_line = max(int(col_width * 1.2), 10)
        lines = 0
        for part in text.split("\n"):
            lines += max(1, (len(part) + chars_per_line - 1) // chars_per_line)
        max_lines = max(max_lines, lines)
    ws.row_dimensions[row_num].height = min(max(min_height, max_lines * 15), max_height)


def _configure_sheet(ws, column_widths: dict) -> None:
    header_font = Font(bold=True)
    for header in ws[1]:
        header.font = header_font
        header.alignment = Alignment(horizontal="center", vertical="center")
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 20


def _apply_multiline_cell(ws, row_num: int, column: int, value: Any) -> None:
    cell = ws.cell(row_num, column, value if value is not None else "")
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_run_export_excel(reports: List[RunOut]) -> BytesIO:
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "报告摘要"
    summary_headers = [
        "报告时间",
        "套件",
        "结果",
        "通过率",
        "通过数",
        "失败数",
        "总数",
        "耗时",
        "触发方式",
        "开始时间",
        "结束时间",
    ]
    summary_ws.append(summary_headers)
    _configure_sheet(
        summary_ws,
        {"A": 20, "B": 24, "C": 10, "D": 10, "E": 8, "F": 8, "G": 8, "H": 10, "I": 12, "J": 20, "K": 20},
    )

    for report in reports:
        row_num = summary_ws.max_row + 1
        summary_ws.cell(row_num, 1, _format_report_time(report.started_at))
        summary_ws.cell(row_num, 2, report.target_name)
        summary_ws.cell(row_num, 3, STATUS_LABELS.get(report.status, report.status))
        summary_ws.cell(row_num, 4, f"{report.pass_rate}%")
        summary_ws.cell(row_num, 5, report.passed_count)
        summary_ws.cell(row_num, 6, report.failed_count)
        summary_ws.cell(row_num, 7, report.total_count)
        summary_ws.cell(row_num, 8, _format_duration(report.duration_ms))
        summary_ws.cell(row_num, 9, report.triggered_by or "")
        summary_ws.cell(row_num, 10, _format_dt(report.started_at))
        summary_ws.cell(row_num, 11, _format_dt(report.finished_at))

    detail_ws = wb.create_sheet("用例明细")
    detail_headers = [
        "报告时间",
        "套件",
        "序号",
        "用例名称",
        "方法",
        "URL",
        "结果",
        "耗时",
        "响应状态",
        "错误信息",
        "断言结果",
        "请求头",
        "请求体",
        "响应头",
        "响应体",
    ]
    detail_ws.append(detail_headers)
    _configure_sheet(
        detail_ws,
        {
            "A": 20, "B": 20, "C": 8, "D": 22, "E": 10, "F": 36, "G": 10, "H": 10,
            "I": 10, "J": 24, "K": 28, "L": 24, "M": 24, "N": 24, "O": 32,
        },
    )

    for report in reports:
        for index, step in enumerate(report.steps, start=1):
            row_num = detail_ws.max_row + 1
            detail_ws.cell(row_num, 1, _format_report_time(report.started_at))
            detail_ws.cell(row_num, 2, report.target_name)
            detail_ws.cell(row_num, 3, index)
            detail_ws.cell(row_num, 4, step.case_name)
            detail_ws.cell(row_num, 5, step.method)
            _apply_multiline_cell(detail_ws, row_num, 6, step.url)
            detail_ws.cell(row_num, 7, STATUS_LABELS.get(step.status, step.status))
            detail_ws.cell(row_num, 8, _format_duration(step.duration_ms))
            detail_ws.cell(row_num, 9, step.response_status if step.response_status is not None else "")
            _apply_multiline_cell(detail_ws, row_num, 10, step.error_message or "")
            _apply_multiline_cell(detail_ws, row_num, 11, _format_assertions(step.assertion_results))
            _apply_multiline_cell(detail_ws, row_num, 12, _pretty_json_text(step.request_headers))
            _apply_multiline_cell(detail_ws, row_num, 13, _pretty_json_text(step.request_body))
            _apply_multiline_cell(detail_ws, row_num, 14, _pretty_json_text(step.response_headers))
            _apply_multiline_cell(detail_ws, row_num, 15, _pretty_json_text(step.response_body))
            _autosize_row(detail_ws, row_num)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

