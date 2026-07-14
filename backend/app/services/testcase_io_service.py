import json
import re
import zipfile
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy.orm import Session, joinedload

from app.models.project import Project
from app.models.requirement import Requirement
from app.models.testcase import TestCase
from app.models.user import User
from app.services.requirement_io_service import (
    GROUP_SUFFIX,
    TYPE_ALIASES,
    TYPE_LABELS,
    _XMIND_THUMBNAIL_PNG,
    _build_xmind_topic,
    _load_xmind_topics,
    _strip_group_label,
    _topic_id,
    normalize_priority,
    normalize_type,
)

CASE_TYPE_LABELS = TYPE_LABELS
CASE_TYPE_ALIASES = TYPE_ALIASES

REVIEW_STATUS_LABELS = {
    "draft": "草稿",
    "pending": "待评审",
    "approved": "已通过",
    "rejected": "已驳回",
}

REVIEW_STATUS_ALIASES = {
    "草稿": "draft",
    "draft": "draft",
    "待评审": "pending",
    "pending": "pending",
    "已通过": "approved",
    "approved": "approved",
    "已驳回": "rejected",
    "rejected": "rejected",
}

SOURCE_LABELS = {
    "manual": "手动",
    "ai_generated": "AI生成",
}

SOURCE_ALIASES = {
    "手动": "manual",
    "manual": "manual",
    "AI生成": "ai_generated",
    "ai_generated": "ai_generated",
}

ALLOWED_CASE_TYPES = set(CASE_TYPE_LABELS)
ALLOWED_PRIORITIES = {"P0", "P1", "P2", "P3"}
ALLOWED_REVIEW_STATUSES = set(REVIEW_STATUS_LABELS)
ALLOWED_SOURCES = set(SOURCE_LABELS)

EXCEL_HEADERS = [
    "ID",
    "需求点",
    "标题",
    "类型",
    "优先级",
    "前置条件",
    "步骤",
    "预期结果",
    "标签",
    "来源",
    "评审状态",
]

HEADER_ALIASES = {
    "id": "ID",
    "需求点": "需求点",
    "requirement": "需求点",
    "标题": "标题",
    "title": "标题",
    "类型": "类型",
    "type": "类型",
    "case_type": "类型",
    "优先级": "优先级",
    "priority": "优先级",
    "前置条件": "前置条件",
    "preconditions": "前置条件",
    "步骤": "步骤",
    "steps": "步骤",
    "预期结果": "预期结果",
    "expected_results": "预期结果",
    "标签": "标签",
    "tags": "标签",
    "来源": "来源",
    "source": "来源",
    "评审状态": "评审状态",
    "review_status": "评审状态",
}

NOTE_SECTIONS = (
    ("前置条件", "preconditions"),
    ("测试步骤", "steps"),
    ("预期结果", "expected_results"),
)


def build_content_disposition(filename: str, fallback: str) -> str:
    encoded = quote(filename, safe="")
    return f'attachment; filename="{fallback}"; filename*=UTF-8\'\'{encoded}'


def normalize_case_type(value: Optional[str]) -> str:
    text = (value or "").strip()
    if not text:
        return "functional"
    mapped = CASE_TYPE_ALIASES.get(text) or CASE_TYPE_ALIASES.get(text.lower())
    if mapped in ALLOWED_CASE_TYPES:
        return mapped
    raise ValueError(f"无效类型: {value}")


def normalize_source(value: Optional[str]) -> str:
    text = (value or "manual").strip()
    mapped = SOURCE_ALIASES.get(text) or SOURCE_ALIASES.get(text.lower())
    if mapped in ALLOWED_SOURCES:
        return mapped
    raise ValueError(f"无效来源: {value}")


def normalize_review_status(value: Optional[str]) -> str:
    text = (value or "draft").strip()
    mapped = REVIEW_STATUS_ALIASES.get(text) or REVIEW_STATUS_ALIASES.get(text.lower())
    if mapped in ALLOWED_REVIEW_STATUSES:
        return mapped
    raise ValueError(f"无效评审状态: {value}")


def _format_numbered_lines(text: Optional[str]) -> Optional[str]:
    if not text:
        return text
    value = str(text).strip()
    if not value:
        return value
    if "\n" in value:
        lines = [line.strip() for line in value.splitlines() if line.strip()]
        return "\n".join(lines)

    value = re.sub(r"[；;]\s*(?=\d+\.)", "\n", value)
    value = re.sub(r"(?<=[^\n])\s+(?=\d+\.)", "\n", value)
    return value.strip()


def _text_line_count(text: Optional[str], column_width: int) -> int:
    content = _format_numbered_lines(text) or ""
    if not content:
        return 1
    total = 0
    for paragraph in content.split("\n"):
        display_width = sum(2 if ord(char) > 127 else 1 for char in paragraph)
        total += max(1, (display_width + column_width - 1) // column_width)
    return max(1, total)


def _apply_multiline_cell_style(ws, row: int, column: int, value: Optional[str]) -> None:
    cell = ws.cell(row, column, _format_numbered_lines(value))
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _autosize_excel_row(ws, row: int) -> None:
    line_counts = [
        _text_line_count(ws.cell(row, 6).value, 28),
        _text_line_count(ws.cell(row, 7).value, 36),
        _text_line_count(ws.cell(row, 8).value, 36),
    ]
    ws.row_dimensions[row].height = 15 * max(line_counts) + 4


def _configure_excel_sheet(ws) -> None:
    column_widths = {
        "A": 8,
        "B": 14,
        "C": 24,
        "D": 10,
        "E": 8,
        "F": 28,
        "G": 36,
        "H": 36,
        "I": 16,
        "J": 12,
        "K": 10,
    }
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 18


def _merge_requirement_column(ws, row_keys: List[Optional[int]], *, column: int = 2) -> None:
    if not row_keys:
        return

    merge_start = 0
    current_key = row_keys[0]
    for index in range(1, len(row_keys)):
        key = row_keys[index]
        if key != current_key or key is None:
            if current_key is not None and index - merge_start > 1:
                start_row = merge_start + 2
                end_row = index + 1
                ws.merge_cells(
                    start_row=start_row,
                    start_column=column,
                    end_row=end_row,
                    end_column=column,
                )
                ws.cell(start_row, column).alignment = Alignment(vertical="center", wrap_text=False)
            merge_start = index
            current_key = key

    if current_key is not None and len(row_keys) - merge_start > 1:
        start_row = merge_start + 2
        end_row = len(row_keys) + 1
        ws.merge_cells(
            start_row=start_row,
            start_column=column,
            end_row=end_row,
            end_column=column,
        )
        ws.cell(start_row, column).alignment = Alignment(vertical="center", wrap_text=False)


def _sort_testcases(cases: List[TestCase]) -> List[TestCase]:
    return sorted(
        cases,
        key=lambda item: (
            item.requirement.title if item.requirement else "",
            item.requirement_id or 0,
            item.id,
        ),
    )


def _build_case_notes(case: TestCase) -> Optional[str]:
    parts: List[str] = []
    if case.preconditions:
        parts.append(f"前置条件：\n{case.preconditions}")
    if case.steps:
        parts.append(f"测试步骤：\n{case.steps}")
    if case.expected_results:
        parts.append(f"预期结果：\n{case.expected_results}")
    if case.tags:
        parts.append(f"标签：{case.tags}")
    return "\n\n".join(parts) if parts else None


def _parse_case_notes(notes: str) -> Dict[str, str]:
    result = {"preconditions": "", "steps": "", "expected_results": "", "tags": ""}
    text = (notes or "").strip()
    if not text:
        return result

    markers = ["前置条件", "测试步骤", "预期结果", "标签"]
    positions: List[Tuple[int, str]] = []
    for marker in markers:
        match = re.search(rf"{marker}[：:]", text)
        if match:
            positions.append((match.start(), marker))

    if not positions:
        result["steps"] = text
        return result

    positions.sort(key=lambda item: item[0])
    for index, (start, marker) in enumerate(positions):
        content_start = start + len(marker) + 1
        content_end = positions[index + 1][0] if index + 1 < len(positions) else len(text)
        content = text[content_start:content_end].strip()
        if marker == "前置条件":
            result["preconditions"] = content
        elif marker == "测试步骤":
            result["steps"] = content
        elif marker == "预期结果":
            result["expected_results"] = content
        elif marker == "标签":
            result["tags"] = content.replace("标签：", "").replace("标签:", "").strip()
    return result


def _infer_requirement_priority(ancestors: List[str]) -> Tuple[str, str]:
    requirement_title = ""
    priority = "P1"
    for label in ancestors:
        clean = _strip_group_label(label)
        if clean.upper() in ALLOWED_PRIORITIES:
            priority = normalize_priority(clean)
        elif clean.startswith("P") and clean[:2] in ALLOWED_PRIORITIES:
            priority = normalize_priority(clean[:2])
        elif clean and clean not in ("未关联需求",):
            requirement_title = clean
    return requirement_title, priority


def _find_requirement_id(
    db: Session,
    project_id: int,
    title: str,
    requirement_map: Optional[Dict[str, int]] = None,
) -> Optional[int]:
    clean = (title or "").strip()
    if not clean or clean in {"未关联需求", "-"}:
        return None
    if requirement_map is not None:
        return requirement_map.get(clean)
    requirement = (
        db.query(Requirement)
        .filter(Requirement.project_id == project_id, Requirement.title == clean)
        .first()
    )
    return requirement.id if requirement else None


def _build_requirement_map(db: Session, project_id: int) -> Dict[str, int]:
    requirements = db.query(Requirement).filter(Requirement.project_id == project_id).all()
    return {req.title.strip(): req.id for req in requirements if req.title}


def list_project_testcases(db: Session, project_id: int) -> List[TestCase]:
    cases = (
        db.query(TestCase)
        .options(joinedload(TestCase.requirement))
        .filter(TestCase.project_id == project_id)
        .all()
    )
    return _sort_testcases(cases)


def export_testcases_excel(project: Project, cases: List[TestCase]) -> Tuple[BytesIO, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(EXCEL_HEADERS)
    for col in range(1, len(EXCEL_HEADERS) + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    _configure_excel_sheet(ws)

    merge_keys: List[Optional[int]] = []
    for case in cases:
        requirement_title = case.requirement.title if case.requirement else ""
        merge_keys.append(case.requirement_id)
        row_num = ws.max_row + 1
        ws.cell(row_num, 1, case.id)
        ws.cell(row_num, 2, requirement_title)
        ws.cell(row_num, 3, case.title)
        ws.cell(row_num, 4, CASE_TYPE_LABELS.get(case.case_type, case.case_type))
        ws.cell(row_num, 5, case.priority)
        _apply_multiline_cell_style(ws, row_num, 6, case.preconditions)
        _apply_multiline_cell_style(ws, row_num, 7, case.steps)
        _apply_multiline_cell_style(ws, row_num, 8, case.expected_results)
        ws.cell(row_num, 9, case.tags or "")
        ws.cell(row_num, 10, SOURCE_LABELS.get(case.source, case.source))
        ws.cell(row_num, 11, REVIEW_STATUS_LABELS.get(case.review_status, case.review_status))
        _autosize_excel_row(ws, row_num)

    _merge_requirement_column(ws, merge_keys, column=2)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{project.name}_testcases.xlsx"
    return buffer, filename


def _group_testcases(cases: List[TestCase]) -> Dict[str, Dict[str, List[TestCase]]]:
    grouped: Dict[str, Dict[str, List[TestCase]]] = {}
    for case in cases:
        req_key = case.requirement.title if case.requirement else "未关联需求"
        priority = case.priority or "P1"
        grouped.setdefault(req_key, {}).setdefault(priority, []).append(case)
    return grouped


def export_testcases_xmind(project: Project, cases: List[TestCase]) -> Tuple[BytesIO, str]:
    grouped = _group_testcases(cases)
    requirement_nodes: List[Dict] = []

    for req_title in sorted(grouped.keys()):
        priority_map = grouped[req_title]
        priority_nodes: List[Dict] = []
        for priority in ["P0", "P1", "P2", "P3"]:
            items = priority_map.get(priority) or []
            if not items:
                continue
            leaf_nodes = [
                _build_xmind_topic(item.title, notes=_build_case_notes(item))
                for item in sorted(items, key=lambda row: row.id, reverse=True)
            ]
            priority_nodes.append(
                _build_xmind_topic(f"{priority}（{len(items)}）", children=leaf_nodes)
            )
        total = sum(len(v) for v in priority_map.values())
        requirement_nodes.append(
            _build_xmind_topic(f"{req_title}（{total}）", children=priority_nodes or None)
        )

    root = _build_xmind_topic(
        f"{project.name}（{len(cases)}）",
        children=requirement_nodes or None,
        structure_class="org.xmind.ui.logic.right",
    )
    sheet_id = _topic_id()
    sheet = {
        "id": sheet_id,
        "class": "sheet",
        "title": "测试用例",
        "rootTopic": root,
        "theme": {
            "map": {
                "id": _topic_id(),
                "properties": {
                    "svg:fill": "#ffffff",
                    "multi-line-colors": "",
                    "color-list": "",
                },
            }
        },
    }
    metadata = {
        "dataStructureVersion": "2",
        "layoutEngineVersion": "3",
        "creator": {"name": "AI质量平台", "version": "1.0.0"},
        "activeSheetId": sheet_id,
        "modifier": "",
    }
    manifest = {
        "file-entries": {
            "content.json": {},
            "metadata.json": {},
            "Thumbnails/thumbnail.png": {},
        }
    }

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("content.json", json.dumps([sheet], ensure_ascii=False))
        archive.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False))
        archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
        archive.writestr("Thumbnails/thumbnail.png", _XMIND_THUMBNAIL_PNG)
    buffer.seek(0)
    filename = f"{project.name}_testcases.xmind"
    return buffer, filename


def _parse_excel_rows(file_bytes: bytes) -> List[Dict[str, str]]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 文件为空")

    header_row = [str(cell or "").strip() for cell in rows[0]]
    mapping: Dict[int, str] = {}
    for index, header in enumerate(header_row):
        normalized = HEADER_ALIASES.get(header) or HEADER_ALIASES.get(header.lower())
        if normalized:
            mapping[index] = normalized

    if "标题" not in mapping.values():
        raise ValueError("Excel 缺少「标题」列")

    items: List[Dict[str, str]] = []
    last_requirement = ""
    for row in rows[1:]:
        if not row:
            continue
        record: Dict[str, str] = {}
        for index, field in mapping.items():
            value = row[index] if index < len(row) else None
            record[field] = str(value).strip() if value is not None else ""
        title = record.get("标题", "").strip()
        if not title:
            continue
        requirement = record.get("需求点", "").strip()
        if requirement:
            last_requirement = requirement
        elif last_requirement:
            record["需求点"] = last_requirement
        record["评审状态"] = "草稿"
        items.append(record)
    if not items:
        raise ValueError("Excel 中没有可导入的用例行")
    return items


def _collect_xmind_testcases(topic: Dict, ancestors: Optional[List[str]] = None) -> List[Dict[str, str]]:
    ancestors = list(ancestors or [])
    title = (topic.get("title") or "").strip()
    children = (topic.get("children") or {}).get("attached") or []
    notes = ((topic.get("notes") or {}).get("plain") or {}).get("content") or ""
    notes = str(notes).strip()

    items: List[Dict[str, str]] = []
    if not children:
        if not title or GROUP_SUFFIX.match(title):
            return items
        if len(ancestors) == 0:
            return items
        clean_title = _strip_group_label(title)
        context = [_strip_group_label(label) for label in ancestors[1:] if label]
        requirement_title, priority = _infer_requirement_priority(context)
        parsed_notes = _parse_case_notes(notes)
        items.append(
            {
                "需求点": requirement_title,
                "标题": clean_title,
                "类型": "功能",
                "优先级": priority,
                "前置条件": parsed_notes["preconditions"],
                "步骤": parsed_notes["steps"],
                "预期结果": parsed_notes["expected_results"],
                "标签": parsed_notes["tags"],
                "来源": "手动",
                "评审状态": "草稿",
            }
        )
        return items

    next_ancestors = ancestors + [title] if title else ancestors
    for child in children:
        items.extend(_collect_xmind_testcases(child, next_ancestors))
    return items


def parse_testcase_import_file(filename: str, file_bytes: bytes) -> List[Dict[str, str]]:
    lower_name = (filename or "").lower()
    if lower_name.endswith((".xlsx", ".xlsm", ".xltx")):
        return _parse_excel_rows(file_bytes)
    if lower_name.endswith(".xmind"):
        topics = _load_xmind_topics(file_bytes)
        items: List[Dict[str, str]] = []
        for topic in topics:
            items.extend(_collect_xmind_testcases(topic, []))
        if not items:
            raise ValueError("XMind 中未找到可导入的用例节点")
        return items
    raise ValueError("仅支持 .xlsx 或 .xmind 文件")


def import_testcases_from_rows(
    db: Session,
    project: Project,
    rows: List[Dict[str, str]],
    current_user: User,
) -> Tuple[int, int]:
    imported = 0
    skipped = 0
    requirement_map = _build_requirement_map(db, project.id)
    for row in rows:
        title = row.get("标题", "").strip()
        if not title:
            skipped += 1
            continue
        try:
            requirement_id = _find_requirement_id(
                db,
                project.id,
                row.get("需求点", ""),
                requirement_map=requirement_map,
            )
            testcase = TestCase(
                project_id=project.id,
                requirement_id=requirement_id,
                title=title,
                case_type=normalize_case_type(row.get("类型")),
                priority=normalize_priority(row.get("优先级")),
                preconditions=row.get("前置条件") or None,
                steps=row.get("步骤") or None,
                expected_results=row.get("预期结果") or None,
                tags=row.get("标签") or None,
                source=normalize_source(row.get("来源")),
                review_status="draft",
                created_by_id=current_user.id,
            )
        except ValueError:
            skipped += 1
            continue
        db.add(testcase)
        imported += 1
    if imported:
        db.commit()
    return imported, skipped
