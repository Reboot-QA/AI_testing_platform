import asyncio
import json
import re
from io import BytesIO
from urllib.parse import quote

from typing import Any, Dict, List, Optional, Union

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from sqlalchemy.orm import Session, joinedload

from app.auth import get_current_user
from app.database import SessionLocal, get_db
from app.models.project import Project
from app.models.requirement import Requirement
from app.models.testcase import TestCase
from app.models.user import User
from app.schemas import (
    AIGenerateRequest,
    AIGenerateResponse,
    BatchDeleteResponse,
    TestCaseBatchDelete,
    TestCaseBatchReviewResponse,
    TestCaseBatchReviewUpdate,
    TestCaseCreate,
    TestCaseOut,
    TestCasePageOut,
    TestCaseUpdate,
)
from app.config import settings
from app.services.ai_service import (
    build_generation_tasks,
    generate_testcases,
    split_generation_batches,
    stream_generate_batches,
)
from app.services.settings_service import get_effective_llm_config

router = APIRouter(prefix="/testcases", tags=["用例"])

ALLOWED_REVIEW_STATUSES = {"draft", "pending", "approved", "rejected"}
REVIEW_TRANSITIONS = {
    "pending": {"draft", "rejected"},
    "approved": {"pending"},
    "rejected": {"pending"},
}


def _check_project(db: Session, project_id: int, user_id: int) -> Project:
    project = db.query(Project).filter(Project.id == project_id, Project.owner_id == user_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    return project


def _build_content_disposition(filename: str) -> str:
    """Build RFC 5987 Content-Disposition header for non-ASCII filenames."""
    ascii_fallback = "testcases.xlsx"
    encoded = quote(filename, safe="")
    return f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{encoded}"


def _prepare_ai_generate(
    db: Session,
    data: AIGenerateRequest,
    current_user: User,
) -> Dict[str, Any]:
    _check_project(db, data.project_id, current_user.id)

    requirement_text = data.requirement_text or ""
    requirement_ids = list(dict.fromkeys(data.requirement_ids or []))
    if data.requirement_id and data.requirement_id not in requirement_ids:
        requirement_ids.insert(0, data.requirement_id)

    selected_requirements: List[Requirement] = []
    if requirement_ids:
        selected_requirements = (
            db.query(Requirement)
            .filter(
                Requirement.id.in_(requirement_ids),
                Requirement.project_id == data.project_id,
            )
            .all()
        )
        if len(selected_requirements) != len(requirement_ids):
            raise HTTPException(status_code=404, detail="部分关联需求不存在")
        not_approved = [req.title for req in selected_requirements if req.status != "approved"]
        if not_approved:
            raise HTTPException(
                status_code=400,
                detail=f"仅已评审需求可关联生成用例：{', '.join(not_approved)}",
            )
        order_map = {item.id: item for item in selected_requirements}
        selected_requirements = [order_map[item_id] for item_id in requirement_ids if item_id in order_map]
        requirement_text = "\n\n".join(
            f"【{req.title}】\n{req.description or ''}" for req in selected_requirements
        )

    if not requirement_text.strip():
        raise HTTPException(status_code=400, detail="请提供需求文本或关联需求")

    llm_config = get_effective_llm_config(db, data.provider_id)
    if data.provider_id and llm_config.get("provider_id") != data.provider_id:
        raise HTTPException(status_code=400, detail="所选模型不存在或已禁用")
    if not llm_config["mock_mode"] and not llm_config["api_key"]:
        raise HTTPException(status_code=400, detail="当前模型未配置 API Key，请前往系统管理配置，或开启 Mock 模式")

    return {
        "requirement_text": requirement_text,
        "requirement_id": selected_requirements[0].id if selected_requirements else None,
        "selected_requirements": selected_requirements,
        "llm_config": llm_config,
    }


def _testcase_out(case: TestCase, db: Session) -> TestCaseOut:
    creator_name = case.creator.username if case.creator else ""
    if not creator_name and case.created_by_id:
        creator = db.query(User).filter(User.id == case.created_by_id).first()
        creator_name = creator.username if creator else ""
    return TestCaseOut(
        id=case.id,
        project_id=case.project_id,
        requirement_id=case.requirement_id,
        title=case.title,
        case_type=case.case_type,
        priority=case.priority,
        preconditions=case.preconditions,
        steps=case.steps,
        expected_results=case.expected_results,
        tags=case.tags,
        source=case.source,
        review_status=case.review_status,
        created_by_id=case.created_by_id,
        creator_name=creator_name,
        created_at=case.created_at,
    )


def _stage_generated_case(
    db: Session,
    *,
    project_id: int,
    requirement_id: Optional[int],
    selected_requirements: List[Requirement],
    case_type: str,
    mode: str,
    item: dict,
    created_by_id: Optional[int] = None,
) -> TestCase:
    case = TestCase(
        project_id=project_id,
        requirement_id=requirement_id,
        title=item.get("title", "未命名用例"),
        case_type=case_type,
        priority=item.get("priority", "P1"),
        preconditions=item.get("preconditions"),
        steps=item.get("steps"),
        expected_results=item.get("expected_results"),
        tags=item.get("tags"),
        source="ai_generated",
        review_status="pending",
        created_by_id=created_by_id,
        ai_metadata=json.dumps(
            {
                "mode": mode,
                "requirement_ids": [req.id for req in selected_requirements],
                "requirement_id": requirement_id,
            },
            ensure_ascii=False,
        ),
    )
    db.add(case)
    db.flush()
    db.refresh(case)
    return case


def _save_generated_case(
    db: Session,
    *,
    project_id: int,
    requirement_id: Optional[int],
    selected_requirements: List[Requirement],
    case_type: str,
    mode: str,
    item: dict,
    created_by_id: Optional[int] = None,
) -> TestCase:
    case = _stage_generated_case(
        db,
        project_id=project_id,
        requirement_id=requirement_id,
        selected_requirements=selected_requirements,
        case_type=case_type,
        mode=mode,
        item=item,
        created_by_id=created_by_id,
    )
    db.commit()
    db.refresh(case)
    return case


def _sse_event(payload: dict) -> str:
    return f"data: {json.dumps(payload, ensure_ascii=False, default=str)}\n\n"


def _build_generation_tasks(ctx: Dict[str, Any], total_count: int) -> List[Dict[str, Any]]:
    req_items = [
        {"id": req.id, "title": req.title, "description": req.description or ""}
        for req in ctx["selected_requirements"]
    ]
    manual_text = ctx["requirement_text"] if not req_items else ""
    return build_generation_tasks(req_items, manual_text, total_count)


@router.get("", response_model=Union[List[TestCaseOut], TestCasePageOut])
def list_testcases(
    project_id: int = Query(...),
    requirement_id: Optional[int] = Query(None),
    review_status: Optional[str] = Query(None),
    page: Optional[int] = Query(None, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _check_project(db, project_id, current_user.id)
    query = (
        db.query(TestCase)
        .options(joinedload(TestCase.creator))
        .filter(TestCase.project_id == project_id)
    )
    if requirement_id:
        query = query.filter(TestCase.requirement_id == requirement_id)
    if review_status:
        query = query.filter(TestCase.review_status == review_status)
    query = query.order_by(TestCase.id.desc())

    if page is not None:
        total = query.count()
        items = query.offset((page - 1) * page_size).limit(page_size).all()
        return TestCasePageOut(
            items=[_testcase_out(item, db) for item in items],
            total=total,
            page=page,
            page_size=page_size,
        )

    return [_testcase_out(item, db) for item in query.all()]


@router.post("", response_model=TestCaseOut)
def create_testcase(
    data: TestCaseCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _check_project(db, data.project_id, current_user.id)
    case = TestCase(**data.model_dump(), created_by_id=current_user.id)
    db.add(case)
    db.commit()
    db.refresh(case)
    return _testcase_out(case, db)


@router.post("/ai/generate", response_model=AIGenerateResponse)
async def ai_generate(
    data: AIGenerateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ctx = _prepare_ai_generate(db, data, current_user)
    llm_config = ctx["llm_config"]
    tasks = _build_generation_tasks(ctx, data.count)

    saved: List[TestCase] = []
    mode = "mock" if llm_config["mock_mode"] else "llm"
    try:
        for task in tasks:
            cases_data, current_mode = await generate_testcases(
                task["requirement_text"],
                data.case_type,
                task["count"],
                api_base=llm_config["api_base"],
                api_key=llm_config["api_key"],
                model=llm_config["model"],
                mock_mode=llm_config["mock_mode"],
            )
            mode = current_mode
            for item in cases_data:
                saved.append(
                    _save_generated_case(
                        db,
                        project_id=data.project_id,
                        requirement_id=task["requirement_id"],
                        selected_requirements=ctx["selected_requirements"],
                        case_type=data.case_type,
                        mode=mode,
                        item=item,
                        created_by_id=current_user.id,
                    )
                )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"AI 生成失败: {exc}") from exc

    return AIGenerateResponse(
        generated_count=len(saved),
        testcases=[_testcase_out(case, db) for case in saved],
        mode=mode,
        provider_name=llm_config.get("provider_name"),
        model=llm_config.get("model"),
    )


@router.post("/ai/generate/stream")
async def ai_generate_stream(
    data: AIGenerateRequest,
    current_user: User = Depends(get_current_user),
):
    db = SessionLocal()
    try:
        ctx = _prepare_ai_generate(db, data, current_user)
    except HTTPException:
        db.close()
        raise
    db.close()

    async def event_generator():
        stream_db = SessionLocal()
        saved_count = 0
        failed_tasks: List[str] = []
        mode = "mock" if ctx["llm_config"]["mock_mode"] else "llm"
        llm_config = ctx["llm_config"]
        tasks = _build_generation_tasks(ctx, data.count)

        try:
            if not tasks:
                yield _sse_event({"type": "error", "message": "没有可生成的需求任务，请检查关联需求"})
                return

            batch_call_count = sum(
                len(split_generation_batches(task["count"], settings.ai_generate_batch_size))
                for task in tasks
            )
            yield _sse_event(
                {
                    "type": "status",
                    "message": (
                        f"正在调用大模型：{len(tasks)} 个需求，"
                        f"{batch_call_count} 批请求（并发 {settings.ai_generate_concurrency}）..."
                    ),
                    "current": 0,
                    "total": data.count,
                }
            )

            async for cases_data, batch_mode, batch_index, batch_total, requirement_id, label, task_error in stream_generate_batches(
                tasks,
                data.case_type,
                api_base=llm_config["api_base"],
                api_key=llm_config["api_key"],
                model=llm_config["model"],
                mock_mode=llm_config["mock_mode"],
                batch_size=settings.ai_generate_batch_size,
                concurrency=settings.ai_generate_concurrency,
            ):
                mode = batch_mode
                if task_error:
                    failed_tasks.append(f"{label}：{task_error}")
                    yield _sse_event(
                        {
                            "type": "status",
                            "message": f"「{label}」生成失败，已跳过（{batch_index}/{batch_total}）",
                            "current": saved_count,
                            "total": data.count,
                        }
                    )
                    continue

                yield _sse_event(
                    {
                        "type": "status",
                        "message": f"正在为「{label}」生成用例（{batch_index}/{batch_total} 批）...",
                        "current": saved_count,
                        "total": data.count,
                    }
                )

                for item in cases_data:
                    case = _stage_generated_case(
                        stream_db,
                        project_id=data.project_id,
                        requirement_id=requirement_id,
                        selected_requirements=ctx["selected_requirements"],
                        case_type=data.case_type,
                        mode=mode,
                        item=item,
                        created_by_id=current_user.id,
                    )
                    stream_db.flush()
                    stream_db.refresh(case)
                    saved_count += 1
                    yield _sse_event(
                        {
                            "type": "case",
                            "data": _testcase_out(case, stream_db).model_dump(mode="json"),
                            "current": saved_count,
                            "total": data.count,
                            "saved": True,
                        }
                    )
                stream_db.commit()

            if saved_count == 0:
                detail = failed_tasks[0] if len(failed_tasks) == 1 else "；".join(failed_tasks[:3])
                if len(failed_tasks) > 3:
                    detail += f" 等 {len(failed_tasks)} 项"
                raise ValueError(detail or "未生成任何用例，请检查 API Key 或开启 Mock 模式")

            message = f"成功生成 {saved_count} 条用例，已实时写入用例库"
            if failed_tasks:
                message += f"，{len(failed_tasks)} 个需求生成失败已跳过"

            yield _sse_event(
                {
                    "type": "done",
                    "mode": mode,
                    "provider_name": llm_config.get("provider_name"),
                    "model": llm_config.get("model"),
                    "generated_count": saved_count,
                    "failed_count": len(failed_tasks),
                    "message": message,
                }
            )
        except Exception as exc:
            stream_db.rollback()
            if saved_count > 0:
                message = f"已写入用例库 {saved_count} 条，后续生成中断：{exc}"
                if failed_tasks:
                    message += f"；{len(failed_tasks)} 个需求生成失败已跳过"
                yield _sse_event(
                    {
                        "type": "done",
                        "mode": mode,
                        "provider_name": llm_config.get("provider_name"),
                        "model": llm_config.get("model"),
                        "generated_count": saved_count,
                        "failed_count": len(failed_tasks),
                        "partial": True,
                        "message": message,
                    }
                )
            else:
                yield _sse_event({"type": "error", "message": f"AI 生成失败: {exc}"})
        finally:
            stream_db.close()

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


def _format_numbered_lines(text: Optional[str]) -> Optional[str]:
    """将「1. …；2. …」或同一行多个编号项转为换行显示。"""
    if not text:
        return text
    value = str(text).strip()
    if not value:
        return value
    if "\n" in value:
        lines = [line.strip() for line in value.splitlines() if line.strip()]
        return "\n".join(lines)

    value = re.sub(r"[；;]\s*(?=\d+\.)", "\n", value)
    value = re.sub(r"(?<=[^\n])\s+(?=\d+\.)", "\n", value)
    return value.strip()


def _text_line_count(text: Optional[str], column_width: int) -> int:
    content = _format_numbered_lines(text) or ""
    if not content:
        return 1
    total = 0
    for paragraph in content.split("\n"):
        display_width = sum(2 if ord(char) > 127 else 1 for char in paragraph)
        total += max(1, (display_width + column_width - 1) // column_width)
    return max(1, total)


def _autosize_excel_row(ws, row: int) -> None:
    line_counts = [
        _text_line_count(ws.cell(row, 6).value, 28),
        _text_line_count(ws.cell(row, 7).value, 36),
        _text_line_count(ws.cell(row, 8).value, 36),
    ]
    ws.row_dimensions[row].height = 15 * max(line_counts) + 4


def _apply_multiline_cell_style(ws, row: int, column: int, value: Optional[str]) -> None:
    cell = ws.cell(row, column, _format_numbered_lines(value))
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _configure_excel_sheet(ws) -> None:
    column_widths = {
        "A": 8,
        "B": 14,
        "C": 24,
        "D": 10,
        "E": 8,
        "F": 28,
        "G": 36,
        "H": 36,
        "I": 16,
        "J": 12,
        "K": 10,
    }
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 18


def _merge_requirement_column(ws, row_keys: List[Optional[int]], *, column: int = 2) -> None:
    """合并需求点列中连续相同 requirement_id 的单元格。"""
    if not row_keys:
        return

    merge_start = 0
    current_key = row_keys[0]
    for index in range(1, len(row_keys)):
        key = row_keys[index]
        if key != current_key or key is None:
            if current_key is not None and index - merge_start > 1:
                start_row = merge_start + 2
                end_row = index + 1
                ws.merge_cells(
                    start_row=start_row,
                    start_column=column,
                    end_row=end_row,
                    end_column=column,
                )
                ws.cell(start_row, column).alignment = Alignment(vertical="center", wrap_text=False)
            merge_start = index
            current_key = key

    if current_key is not None and len(row_keys) - merge_start > 1:
        start_row = merge_start + 2
        end_row = len(row_keys) + 1
        ws.merge_cells(
            start_row=start_row,
            start_column=column,
            end_row=end_row,
            end_column=column,
        )
        ws.cell(start_row, column).alignment = Alignment(vertical="center", wrap_text=False)


@router.get("/export/excel")
def export_excel(
    project_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    project = _check_project(db, project_id, current_user.id)
    cases = (
        db.query(TestCase)
        .options(joinedload(TestCase.requirement))
        .filter(TestCase.project_id == project_id)
        .all()
    )
    cases.sort(
        key=lambda item: (
            item.requirement.title if item.requirement else "",
            item.requirement_id or 0,
            item.id,
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    headers = ["ID", "需求点", "标题", "类型", "优先级", "前置条件", "步骤", "预期结果", "标签", "来源", "评审状态"]
    ws.append(headers)
    _configure_excel_sheet(ws)

    merge_keys: List[Optional[int]] = []
    for case in cases:
        requirement_title = case.requirement.title if case.requirement else ""
        merge_keys.append(case.requirement_id)
        row_num = ws.max_row + 1
        ws.cell(row_num, 1, case.id)
        ws.cell(row_num, 2, requirement_title)
        ws.cell(row_num, 3, case.title)
        ws.cell(row_num, 4, case.case_type)
        ws.cell(row_num, 5, case.priority)
        _apply_multiline_cell_style(ws, row_num, 6, case.preconditions)
        _apply_multiline_cell_style(ws, row_num, 7, case.steps)
        _apply_multiline_cell_style(ws, row_num, 8, case.expected_results)
        ws.cell(row_num, 9, case.tags)
        ws.cell(row_num, 10, case.source)
        ws.cell(row_num, 11, case.review_status)
        _autosize_excel_row(ws, row_num)

    _merge_requirement_column(ws, merge_keys, column=2)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{project.name}_testcases.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _build_content_disposition(filename)},
    )


@router.post("/batch/review", response_model=TestCaseBatchReviewResponse)
def batch_review_testcases(
    data: TestCaseBatchReviewUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if data.review_status not in ALLOWED_REVIEW_STATUSES:
        raise HTTPException(status_code=400, detail="无效的评审状态")

    _check_project(db, data.project_id, current_user.id)
    cases = db.query(TestCase).filter(
        TestCase.id.in_(data.case_ids),
        TestCase.project_id == data.project_id,
    ).all()
    if len(cases) != len(set(data.case_ids)):
        raise HTTPException(status_code=404, detail="部分用例不存在")

    allowed_from = REVIEW_TRANSITIONS.get(data.review_status)
    updated = 0
    skipped = 0
    for case in cases:
        if allowed_from and case.review_status not in allowed_from:
            skipped += 1
            continue
        case.review_status = data.review_status
        updated += 1

    if updated == 0:
        status_labels = {
            "pending": "提交评审",
            "approved": "通过",
            "rejected": "驳回",
        }
        raise HTTPException(
            status_code=400,
            detail=f"所选用例均无法批量{status_labels.get(data.review_status, '更新')}",
        )

    db.commit()
    message = f"成功{('通过' if data.review_status == 'approved' else '驳回' if data.review_status == 'rejected' else '提交评审')} {updated} 条用例"
    if skipped:
        message += f"，跳过 {skipped} 条状态不符的用例"
    return TestCaseBatchReviewResponse(
        updated_count=updated,
        skipped_count=skipped,
        message=message,
    )


@router.post("/batch/delete", response_model=BatchDeleteResponse)
def batch_delete_testcases(
    data: TestCaseBatchDelete,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _check_project(db, data.project_id, current_user.id)
    cases = db.query(TestCase).filter(
        TestCase.id.in_(data.case_ids),
        TestCase.project_id == data.project_id,
    ).all()
    if not cases:
        raise HTTPException(status_code=404, detail="未找到可删除的用例")
    for case in cases:
        db.delete(case)
    db.commit()
    return BatchDeleteResponse(deleted_count=len(cases), message=f"成功删除 {len(cases)} 条用例")


@router.get("/{case_id}", response_model=TestCaseOut)
def get_testcase(case_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    case = db.query(TestCase).options(joinedload(TestCase.creator)).filter(TestCase.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404, detail="用例不存在")
    _check_project(db, case.project_id, current_user.id)
    return _testcase_out(case, db)


@router.put("/{case_id}", response_model=TestCaseOut)
def update_testcase(
    case_id: int,
    data: TestCaseUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    case = db.query(TestCase).filter(TestCase.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404, detail="用例不存在")
    _check_project(db, case.project_id, current_user.id)
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(case, key, value)
    db.commit()
    db.refresh(case)
    return _testcase_out(case, db)


@router.delete("/{case_id}")
def delete_testcase(case_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    case = db.query(TestCase).filter(TestCase.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404, detail="用例不存在")
    _check_project(db, case.project_id, current_user.id)
    db.delete(case)
    db.commit()
    return {"message": "删除成功"}
